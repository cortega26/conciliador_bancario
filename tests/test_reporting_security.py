from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from conciliador_bancario.cli import app


def test_reporte_previene_excel_injection(tmp_path: Path) -> None:
    runner = CliRunner()
    out = tmp_path / "out"
    out.mkdir()

    res = runner.invoke(
        app,
        [
            "run",
            "--config",
            str(Path("examples") / "config_cliente.yaml"),
            "--bank",
            str(Path("tests/golden/datasets/csv/banco_sucio.csv")),
            "--expected",
            str(Path("tests/golden/datasets/csv/esperados_sucio.csv")),
            "--out",
            str(out),
        ],
    )
    assert res.exit_code == 0, res.stdout

    wb = load_workbook(out / "reporte_conciliacion.xlsx", read_only=True, data_only=True)
    ws = wb["Transacciones"]
    # Busca cualquier celda de descripcion que empiece con apostrofe (sanitizado).
    # Columna "descripcion" esta definida por el header en excel_report.py.
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    desc_idx = headers.index("descripcion") + 1
    values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values.append(row[desc_idx - 1])
    assert any(isinstance(v, str) and v.startswith("'=") for v in values)
