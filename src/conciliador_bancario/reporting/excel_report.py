from __future__ import annotations

import json
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from conciliador_bancario.models import ConfiguracionCliente, ResultadoConciliacion
from conciliador_bancario.utils.masking import enmascarar_texto_sensible, prevenir_csv_injection


def _ws_write_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _mask_cell(v: object, *, mask: bool) -> object:
    if v is None:
        return ""
    if isinstance(v, (int, float, Decimal)):
        return v
    if isinstance(v, str):
        s = enmascarar_texto_sensible(v) if mask else v
        s = prevenir_csv_injection(s)
        return s
    return v


def generar_reporte_excel(
    path: Path, resultado: ResultadoConciliacion, *, mask: bool, cfg: ConfiguracionCliente
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Intento de determinismo: metadatos fijos. (openpyxl puede sobreescribir `modified` al guardar)
    fixed = datetime(2000, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    wb.properties.created = fixed
    wb.properties.modified = fixed

    # Orden estable
    txs = sorted(resultado.transacciones_bancarias, key=lambda t: t.id)
    exps = sorted(resultado.movimientos_esperados, key=lambda e: e.id)
    matches = sorted(resultado.matches, key=lambda m: m.id)
    hallazgos = sorted(resultado.hallazgos, key=lambda h: h.id)

    # Resumen tecnico
    ws = wb.create_sheet("Resumen")
    ws.append(["Campo", "Valor"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for k, v in [
        ("Run ID", resultado.run_id),
        ("Cliente", cfg.cliente),
        ("Transacciones (banco)", len(txs)),
        ("Movimientos (esperados)", len(exps)),
        ("Matches", len(matches)),
        ("Hallazgos", len(hallazgos)),
        ("Mask", bool(mask)),
    ]:
        ws.append([k, _mask_cell(v, mask=mask)])
    ws.freeze_panes = "A2"

    # Transacciones (banco)
    ws_tx = wb.create_sheet("Transacciones")
    tx_headers = [
        "tx_id",
        "fecha_operacion",
        "fecha_contable",
        "monto",
        "moneda",
        "descripcion",
        "referencia",
        "origen",
        "archivo_origen",
        "fila_origen",
        "cuenta_mask",
        "banco",
        "bloquea_autoconcilia",
        "motivo_bloqueo",
    ]
    tx_rows: list[list[object]] = []
    for tx in txs:
        tx_rows.append(
            [
                tx.id,
                str(tx.fecha_operacion.valor),
                str(tx.fecha_contable.valor) if tx.fecha_contable is not None and tx.fecha_contable.valor else "",
                str(tx.monto.valor),
                tx.moneda,
                _mask_cell(str(tx.descripcion.valor), mask=mask),
                _mask_cell(str(tx.referencia.valor), mask=mask) if tx.referencia is not None else "",
                tx.origen.value,
                _mask_cell(tx.archivo_origen, mask=mask),
                tx.fila_origen or "",
                _mask_cell(tx.cuenta_mask or "", mask=mask),
                _mask_cell(tx.banco or "", mask=mask),
                bool(tx.bloquea_autoconcilia),
                _mask_cell(tx.motivo_bloqueo_autoconcilia or "", mask=mask),
            ]
        )
    _ws_write_table(ws_tx, tx_headers, tx_rows)

    # Esperados
    ws_exp = wb.create_sheet("Esperados")
    exp_headers = ["exp_id", "fecha", "monto", "moneda", "descripcion", "referencia", "tercero"]
    exp_rows: list[list[object]] = []
    for e in exps:
        exp_rows.append(
            [
                e.id,
                str(e.fecha.valor),
                str(e.monto.valor),
                e.moneda,
                _mask_cell(str(e.descripcion.valor), mask=mask),
                _mask_cell(str(e.referencia.valor), mask=mask) if e.referencia is not None else "",
                _mask_cell(str(e.tercero.valor), mask=mask) if e.tercero is not None else "",
            ]
        )
    _ws_write_table(ws_exp, exp_headers, exp_rows)

    # Matches
    ws_m = wb.create_sheet("Matches")
    m_headers = [
        "match_id",
        "estado",
        "score",
        "regla",
        "tx_ids",
        "exp_ids",
        "bloqueado_por_confianza",
        "explicacion",
    ]
    m_rows: list[list[object]] = []
    for m in matches:
        m_rows.append(
            [
                m.id,
                m.estado.value,
                float(m.score),
                m.regla,
                ",".join(m.transacciones_bancarias),
                ",".join(m.movimientos_esperados),
                bool(m.bloqueado_por_confianza),
                _mask_cell(m.explicacion, mask=mask),
            ]
        )
    _ws_write_table(ws_m, m_headers, m_rows)

    # Hallazgos
    ws_h = wb.create_sheet("Hallazgos")
    h_headers = ["hallazgo_id", "severidad", "tipo", "mensaje", "entidad", "entidad_id", "detalles_json"]
    h_rows: list[list[object]] = []
    for h in hallazgos:
        detalles = json.dumps(h.detalles or {}, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
        h_rows.append(
            [
                h.id,
                h.severidad.value,
                h.tipo,
                _mask_cell(h.mensaje, mask=mask),
                h.entidad,
                h.entidad_id or "",
                _mask_cell(detalles, mask=mask),
            ]
        )
    _ws_write_table(ws_h, h_headers, h_rows)

    wb.save(path)

