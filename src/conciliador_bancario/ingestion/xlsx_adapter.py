from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from conciliador_bancario.audit.audit_log import AuditEvent, JsonlAuditWriter
from conciliador_bancario.ingestion.base import ErrorIngestion
from conciliador_bancario.models import (
    CampoConConfianza,
    ConfiguracionCliente,
    MetadataConfianza,
    MovimientoEsperado,
    NivelConfianza,
    OrigenDato,
    TransaccionBancaria,
)
from conciliador_bancario.utils.hashing import sha256_json_estable
from conciliador_bancario.utils.masking import enmascarar_cuenta
from conciliador_bancario.utils.parsing import (
    ErrorParseo,
    normalizar_referencia,
    normalizar_texto,
    parse_fecha_chile,
    parse_monto_clp,
)


def _campo(valor: Any, *, origen: OrigenDato, degrade: float = 0.0) -> CampoConConfianza:
    base = 0.90
    score = max(0.0, min(1.0, base - degrade))
    nivel = (
        NivelConfianza.alta
        if score >= 0.85
        else (NivelConfianza.media if score >= 0.55 else NivelConfianza.baja)
    )
    return CampoConConfianza(valor=valor, confianza=MetadataConfianza(score=score, nivel=nivel, origen=origen))


def _id(path: Path, fila: int, data_norm: dict[str, Any], prefix: str) -> str:
    return f"{prefix}-" + sha256_json_estable({"file": path.name, "row": fila, "data": data_norm})[:12]


def _as_text(v: Any) -> str:
    return "" if v is None else str(v)


def _as_date(v: Any) -> date:
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    return parse_fecha_chile(_as_text(v))


def _keynorm(s: str) -> str:
    return normalizar_texto(s).lower().replace(" ", "_")


def _build_header_map(header: tuple[Any, ...] | None) -> dict[str, int]:
    if not header:
        return {}
    return {_keynorm(str(h)): idx for idx, h in enumerate(header) if h is not None}


def _select_worksheet_with_columns(wb, required: list[list[str]]) -> tuple[Any, dict[str, int], tuple[Any, ...]]:
    """
    required: lista de grupos; cada grupo contiene alias aceptados.
    Se selecciona la primera hoja que contenga al menos 1 alias por grupo.
    """
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        header_map = _build_header_map(header)
        ok = True
        for group in required:
            if not any(a in header_map for a in group):
                ok = False
                break
        if ok:
            return ws, header_map, header
    raise ErrorIngestion("XLSX: no se encontro una hoja con las columnas requeridas.")


def cargar_transacciones_xlsx(
    path: Path, *, cfg: ConfiguracionCliente, audit: JsonlAuditWriter
) -> list[TransaccionBancaria]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws, header_map, header = _select_worksheet_with_columns(
        wb,
        required=[
            ["fecha_operacion", "fecha", "fecha_movimiento"],
            ["monto", "importe", "valor"],
            ["descripcion", "glosa", "detalle", "concepto"],
        ],
    )

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_fecha_op = col("fecha_operacion", "fecha", "fecha_movimiento")
    c_fecha_ct = col("fecha_contable", "fecha_valor", "fecha_proceso")
    c_monto = col("monto", "importe", "valor")
    c_moneda = col("moneda", "currency")
    c_desc = col("descripcion", "glosa", "detalle", "concepto")
    c_ref = col("referencia", "ref", "comprobante", "folio", "nro_referencia")
    c_cuenta = col("cuenta", "nro_cuenta", "cuenta_origen")

    audit.write(AuditEvent("ingestion", "XLSX banco cargado", {"archivo": path.name, "hoja": ws.title}))

    out: list[TransaccionBancaria] = []
    rows = ws.iter_rows(values_only=True)
    _ = next(rows, None)  # header
    for excel_row_idx, row in enumerate(rows, start=2):
        if not row or not any(v is not None and str(v).strip() for v in row):
            continue
        try:
            fecha_op = _as_date(row[c_fecha_op])
            fecha_ct = _as_date(row[c_fecha_ct]) if (c_fecha_ct is not None and row[c_fecha_ct]) else None
            monto: Decimal = parse_monto_clp(_as_text(row[c_monto]))
        except (ErrorParseo, IndexError) as e:
            raise ErrorIngestion(f"Fila {excel_row_idx}: parseo invalido: {e}") from e

        moneda = (normalizar_texto(_as_text(row[c_moneda]) if c_moneda is not None else "") or cfg.moneda_default).upper()
        desc = normalizar_texto(_as_text(row[c_desc]))
        ref_raw = normalizar_texto(_as_text(row[c_ref]) if c_ref is not None else "")
        ref = normalizar_referencia(ref_raw) if ref_raw else ""
        cuenta_raw = normalizar_texto(_as_text(row[c_cuenta]) if c_cuenta is not None else "")

        data_norm = {
            "fecha_operacion": str(fecha_op),
            "fecha_contable": str(fecha_ct) if fecha_ct else None,
            "monto": str(monto),
            "moneda": moneda,
            "descripcion": desc,
            "referencia": ref or None,
        }
        tx_id = _id(path, excel_row_idx, data_norm, "TX")

        origen = OrigenDato.xlsx
        out.append(
            TransaccionBancaria(
                id=tx_id,
                cuenta_mask=enmascarar_cuenta(cuenta_raw) if cuenta_raw else None,
                bloquea_autoconcilia=False,
                motivo_bloqueo_autoconcilia=None,
                fecha_operacion=_campo(fecha_op, origen=origen),
                fecha_contable=_campo(fecha_ct, origen=origen, degrade=0.10) if fecha_ct else None,
                monto=_campo(monto, origen=origen),
                moneda=moneda,
                descripcion=_campo(desc, origen=origen),
                referencia=_campo(ref, origen=origen) if ref else None,
                archivo_origen=path.name,
                origen=origen,
                fila_origen=excel_row_idx,
            )
        )
    return out


def cargar_movimientos_esperados_xlsx(
    path: Path, *, cfg: ConfiguracionCliente, audit: JsonlAuditWriter
) -> list[MovimientoEsperado]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws, header_map, header = _select_worksheet_with_columns(
        wb,
        required=[
            ["fecha", "fecha_documento"],
            ["monto", "importe", "valor"],
            ["descripcion", "glosa", "detalle", "concepto"],
        ],
    )

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_id = col("id", "id_externo")
    c_fecha = col("fecha", "fecha_documento")
    c_monto = col("monto", "importe", "valor")
    c_moneda = col("moneda", "currency")
    c_desc = col("descripcion", "glosa", "detalle", "concepto")
    c_ref = col("referencia", "ref", "folio", "nro_referencia")
    c_terc = col("tercero", "proveedor", "cliente")

    audit.write(AuditEvent("ingestion", "XLSX esperados cargado", {"archivo": path.name, "hoja": ws.title}))

    out: list[MovimientoEsperado] = []
    rows = ws.iter_rows(values_only=True)
    _ = next(rows, None)  # header
    for excel_row_idx, row in enumerate(rows, start=2):
        if not row or not any(v is not None and str(v).strip() for v in row):
            continue
        try:
            fecha = _as_date(row[c_fecha])
            monto: Decimal = parse_monto_clp(_as_text(row[c_monto]))
        except (ErrorParseo, IndexError) as e:
            raise ErrorIngestion(f"Fila {excel_row_idx}: parseo invalido: {e}") from e

        moneda = (normalizar_texto(_as_text(row[c_moneda]) if c_moneda is not None else "") or cfg.moneda_default).upper()
        desc = normalizar_texto(_as_text(row[c_desc]))
        ref_raw = normalizar_texto(_as_text(row[c_ref]) if c_ref is not None else "")
        ref = normalizar_referencia(ref_raw) if ref_raw else ""
        tercero = normalizar_texto(_as_text(row[c_terc]) if c_terc is not None else "")
        id_ext = normalizar_texto(_as_text(row[c_id]) if c_id is not None else "")

        data_norm = {
            "fecha": str(fecha),
            "monto": str(monto),
            "moneda": moneda,
            "descripcion": desc,
            "referencia": ref or None,
            "tercero": tercero or None,
        }
        exp_id = id_ext if id_ext else _id(path, excel_row_idx, data_norm, "EXP")

        origen = OrigenDato.xlsx
        out.append(
            MovimientoEsperado(
                id=exp_id,
                fecha=_campo(fecha, origen=origen),
                monto=_campo(monto, origen=origen),
                moneda=moneda,
                descripcion=_campo(desc, origen=origen),
                referencia=_campo(ref, origen=origen) if ref else None,
                tercero=_campo(tercero, origen=origen, degrade=0.20) if tercero else None,
            )
        )
    return out
